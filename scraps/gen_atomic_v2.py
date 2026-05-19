"""Generate new ATOMIC_CRITERIA_L3 từ format chính thức AUTOMIC-L3-PHAN-TAN-DUNG.xlsx
   Định dạng mới:
   - 1 sheet duy nhất (tên = "P.T. DŨNG 2021" — tên bé + năm sinh)
   - Header row 2: Câu | Mã | Tiêu chí atomic | Trọng số | Mức (NHẬP) | Điểm đạt |
                   Nhóm Năng Lực | Năng lực | Mức độ hỗ trợ | Cấp Bloom | Mức độ khó
   - Hệ số tuyến tính: 0→0 | 1→0.2 | 2→0.4 | 3→0.6 | 4→0.8 | 5→1.0
   - 61 atomic, tổng max 160 (12*10 + 2*20)
"""
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'C:/Users/Admin/Downloads/AUTOMIC - L3 - PHAN TAN DUNG.xlsx', data_only=True)
sheet = wb[wb.sheetnames[0]]

ABILITY_MAP = {
    'Chú ý':'attention', 'Quan sát':'observation', 'Ghi nhớ':'memory',
    'Tập trung':'attention',
    'Số & tính toán':'number', 'Hình học':'geometry', 'Hình học KG':'geometry',
    'Đo lường':'measurement', 'Kiểu mẫu':'pattern', 'Dữ liệu':'data',
    'Hiểu biết':'understanding', 'Hiểu':'understanding',
    'Ứng dụng':'application', 'Phân tích':'analysis', 'Tổng hợp':'synthesis',
    'Trôi chảy':'fluency', 'Linh hoạt':'flexibility',
    'Độc đáo':'originality', 'Chính xác':'precision',
}

criteria = []
for r, row in enumerate(sheet.iter_rows(values_only=True), 1):
    if r < 3: continue
    cau, ma, label, weight, muc, diem_dat, nhom, nang_luc = row[:8]
    if cau is None or not isinstance(ma, str): continue
    if 'Tổng' in str(cau): continue
    if not ma or '.' not in str(ma): continue
    skill = ABILITY_MAP.get(str(nang_luc).strip(), None)
    if not skill:
        print('UNKNOWN skill:', repr(nang_luc), 'for', ma)
        skill = 'observation'
    bloom = row[9] if len(row) > 9 else ''
    difficulty = row[10] if len(row) > 10 else ''
    support = row[8] if len(row) > 8 else ''
    criteria.append({
        'cau': int(cau),
        'ma': str(ma).strip(),
        'label': str(label).strip(),
        'weight': float(weight),
        'skill': skill,
        'nhom': str(nhom).strip() if nhom else '',
        'bloom': str(bloom).strip() if bloom else '',
        'difficulty': str(difficulty).strip() if difficulty else '',
        'support': int(support) if isinstance(support, (int, float)) else 0,
    })

print(f'Total criteria: {len(criteria)}')
total_weight = sum(c['weight'] for c in criteria)
print(f'Total weight: {total_weight}')
# Stats
from collections import Counter
print('Per câu:')
per_cau = Counter()
weight_per_cau = {}
for c in criteria:
    per_cau[c['cau']] += 1
    weight_per_cau[c['cau']] = weight_per_cau.get(c['cau'], 0) + c['weight']
for cau in sorted(per_cau):
    print(f"  Câu {cau}: {per_cau[cau]} items, max {weight_per_cau[cau]}")

# Save JSON
with open(r'C:/Users/Admin/Downloads/wks/scraps/atomic_l3_v2.json', 'w', encoding='utf-8') as f:
    json.dump(criteria, f, ensure_ascii=False, indent=2)

# Generate JS
def jss(s):
    s = str(s).replace(chr(92), chr(92)+chr(92)).replace(chr(34), chr(92)+chr(34))
    return chr(34) + s + chr(34)

lines = []
lines.append('/* ═══════════════════════════════════════════════════════════════')
lines.append('   ATOMIC SCORING L3 v2 — 61 tiêu chí, hệ số tuyến tính 6 mức')
lines.append('   Source: AUTOMIC-L3-PHAN-TAN-DUNG.xlsx (final format)')
lines.append('   Formula: điểm đạt = trọng số × hệ số nhân')
lines.append('     5→1.00 | 4→0.80 | 3→0.60 | 2→0.40 | 1→0.20 | 0→0.00')
lines.append('   Tổng max: 160 điểm — 14 câu (12×10 + 2×20 cho câu 13,14)')
lines.append('   Cột bổ sung: nhóm, bloom, difficulty, support (1-4)')
lines.append('   ═══════════════════════════════════════════════════════════════ */')
lines.append('const __ATOMIC_MULTIPLIER_L3 = {5:1.0, 4:0.8, 3:0.6, 2:0.4, 1:0.2, 0:0.0};')
lines.append('const __ATOMIC_LEVELS_L3 = [5, 4, 3, 2, 1, 0];  // thứ tự hiển thị UI')
lines.append('')
lines.append('const __ATOMIC_CRITERIA_L3 = [')
for c in criteria:
    parts = [
        f"cau:{c['cau']}",
        f"ma:{jss(c['ma'])}",
        f"label:{jss(c['label'])}",
        f"weight:{c['weight']}",
        f"skill:{jss(c['skill'])}",
        f"nhom:{jss(c['nhom'])}",
        f"bloom:{jss(c['bloom'])}",
        f"difficulty:{jss(c['difficulty'])}",
        f"support:{c['support']}",
    ]
    lines.append('  {' + ', '.join(parts) + '},')
lines.append('];')

with open(r'C:/Users/Admin/Downloads/wks/scraps/atomic_l3_v2_data.js', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('Saved scraps/atomic_l3_v2_data.js')
